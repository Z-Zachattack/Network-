import re
from html import unescape
from openpyxl import Workbook

def extract_questions_and_answers(xhtml_text):
    questions_and_answers = re.findall(r'<li id="(.*?)-ex-\d+" class="calibre8">(.*?)<ol class="upper-alpha">(.*?)</ol>', xhtml_text, re.DOTALL)
    
    extracted_data = []
    
    for section, question, answer_options in questions_and_answers:
        # Remove HTML tags and unescape entities from the question
        cleaned_question = unescape(re.sub('<[^<]+?>', '', question)).strip()
        
        # Process options
        options = re.findall(r'<li class="calibre8">(.*?)</li>', answer_options, re.DOTALL)
        cleaned_options = [unescape(re.sub('<[^<]+?>', '', option)).strip() for option in options]
        
        extracted_data.append({
            'section': section.upper(),
            'question': cleaned_question,
            'options': cleaned_options
        })
    
    return extracted_data

def create_excel_file(data):
    # Create a new Workbook
    wb = Workbook()
    ws = wb.active
    
    # Set headers
    ws['A1'] = 'Section'
    ws['B1'] = 'ID'
    ws['C1'] = 'Question'
    ws['D1'] = 'A'
    ws['E1'] = 'B'
    ws['F1'] = 'C'
    ws['G1'] = 'D'
    ws['H1'] = 'E'
    ws['I1'] = 'F'
    ws['J1'] = 'G'
    
    # Populate Excel file with data
    row_num = 2
    for item in data:
        ws[f'A{row_num}'] = item['section']
        ws[f'C{row_num}'] = item['question']
        
        for col_num, option in enumerate(item['options'], start=4):  # Start enumeration from 1 for column D
            col_letter = chr(ord('A') + col_num - 1)
            ws[f'{col_letter}{row_num}'] = option
        
        row_num += 1
    
    # Save the Excel file
    wb.save('questions.xlsx')

# Read XHTML content from a file
with open('network+questions.xhtml', 'r', encoding='utf-8') as file:
    xhtml_text = file.read()

# Extract questions and answers
quiz_data = extract_questions_and_answers(xhtml_text)

# Create Excel file
create_excel_file(quiz_data)