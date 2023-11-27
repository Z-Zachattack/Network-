import re
from html import unescape
from openpyxl import Workbook

def remove_span_content(xhtml_text):
    return re.sub(r'<span.*?</span>', '', xhtml_text, flags=re.DOTALL)

def extract_letters_and_block_text(xhtml_text):
    items = re.findall(r'<li id=".*?ex-(\d+)".*?>([A-Z, ]+)\. (.*?)</li>', xhtml_text, re.DOTALL)
    
    extracted_data = []
    
    for item_id, letters, block_text in items:
        # Extract the number at the end of the ID
        extracted_id = item_id.strip()

        # Remove spaces and split letters into a list
        letter_list = [letter.strip() for letter in letters.replace(' ', '')]

        # Clean the block text by removing HTML tags and unescaping entities
        cleaned_block_text = unescape(re.sub('<[^<]+?>', '', block_text)).strip()

        extracted_data.append({
            'id': extracted_id,
            'letters': letter_list,
            'block_text': cleaned_block_text
        })
    
    # Sort the data based on the ID
    extracted_data.sort(key=lambda x: x['id'])
    
    return extracted_data

def create_excel_file(data):
    # Create a new Workbook
    wb = Workbook()
    ws = wb.active
    
    # Set headers
    ws['A1'] = 'ID'
    ws['B1'] = 'Block Text'
    
    # Populate Excel file with headers for letters
    max_letters = max(len(item['letters']) for item in data)
    for i in range(max_letters):
        ws.cell(row=1, column=i + 3, value=f'Letter {i + 1}')
    
    # Populate Excel file with data
    row_num = 2
    for item in data:
        ws[f'A{row_num}'] = item['id']
        ws[f'B{row_num}'] = item['block_text']
        
        # Populate letters in separate columns without comma
        letters_concatenated = ''.join(item['letters']).replace(',', '')
        for i, letter in enumerate(letters_concatenated, start=1):
            ws.cell(row=row_num, column=i + 2, value=letter)
        
        row_num += 1
    
    # Save the Excel file
    wb.save('answers.xlsx')

# Read XHTML content from a file
with open('answers.xhtml', 'r', encoding='utf-8') as file:
    xhtml_text = file.read()

# Remove content between <span> and </span>
xhtml_text = remove_span_content(xhtml_text)

# Extract letters and block text
letters_and_block_text_data = extract_letters_and_block_text(xhtml_text)

# Create Excel file
create_excel_file(letters_and_block_text_data)

