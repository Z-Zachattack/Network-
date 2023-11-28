from flask import Flask, render_template
import openpyxl
import random

app = Flask(__name__)

# Define the custom filter function
@app.template_filter('index_to_letter')
def index_to_letter(index):
    if 0 <= index < 26:
        return chr(ord('A') + index)
    else:
        return ''
def load_quiz_data(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws_questions = wb['Questions']
    ws_answers = wb['Answers']

    quiz_data = []
    for row_questions, row_answers in zip(ws_questions.iter_rows(min_row=2, values_only=True), ws_answers.iter_rows(min_row=2, values_only=True)):
        section, id, question, *options = row_questions
        correct_answer_indices = [index for index, value in enumerate(row_answers[1:], start=0) if value]
        correct_answer = [index_to_letter(index) for index in correct_answer_indices]
        block_text = list(row_answers[0])
        quiz_data.append({'section': section, 'id': id, 'question': question, 'options': options, 'correct_answer': correct_answer, 'correct_answer_indices': correct_answer_indices, 'block_text': block_text})

    return quiz_data


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/quiz')
def quiz():
    quiz_data = load_quiz_data('quiz.xlsx')
    random_question = random.choice(quiz_data)

    chapter_mapping = {
        'C01': 'Chapter 01 - Networking Fundamentals',
        'C02': 'Chapter 02 - Network Implementations',
        'C03': 'Chapter 03 - Network Operations',
        'C04': 'Chapter 04 - Network Security',
        'C05': 'Chapter 05 - Network Troubleshooting',
        'C06': 'Chapter 06 - Practice Exam 1',
        'C07': 'Chapter 07 - Practice Exam 2',
    }
    id = random_question['id']
    return render_template('quiz.html', id=id, question=random_question, chapter_mapping=chapter_mapping)

if __name__ == '__main__':
    app.run(debug=True)
